"""Exportacion e importacion de Excel, incluida la transaccionalidad."""
import pytest
from openpyxl import load_workbook

from services import activity_service, excel_service, grade_service

USER = {"id": 1, "full_name": "Carlos Profe", "role": "teacher"}


@pytest.fixture()
def escenario(app, data):
    with app.app_context():
        actividad = activity_service.create_activity(
            data["assignment_id"], data["period_id"], "Taller 1", "taller", 40)
        grade_service.save_grade(actividad, data["student_id"], 3.0, USER)
        return {"activity_id": actividad, **data}


def _assignment(app, data):
    from services.academic_service import list_assignments
    with app.app_context():
        return next(a for a in list_assignments()
                    if a["id"] == data["assignment_id"])


def _period(app, data):
    from services.academic_service import get_period
    with app.app_context():
        return get_period(data["period_id"])


def test_exportar_genera_un_archivo_legible(app, escenario):
    with app.app_context():
        buffer = excel_service.export_gradebook(
            _assignment(app, escenario), _period(app, escenario))
    book = load_workbook(buffer)
    sheet = book["Notas"]
    encabezados = [cell.value for cell in sheet[4]]
    assert encabezados[0] == "Codigo"
    assert "Taller 1 (40%)" in encabezados
    codigos = [sheet.cell(row=fila, column=1).value for fila in (5, 6)]
    assert "EST001" in codigos


def test_importar_previsualiza_sin_guardar(app, escenario):
    with app.app_context():
        buffer = excel_service.export_gradebook(
            _assignment(app, escenario), _period(app, escenario))
        book = load_workbook(buffer)
        sheet = book["Notas"]
        sheet.cell(row=5, column=3, value=4.8)   # cambia la nota del primer alumno
        from io import BytesIO
        modificado = BytesIO()
        book.save(modificado)
        modificado.seek(0)

        preview = excel_service.preview_import(
            modificado, escenario["assignment_id"], escenario["period_id"])

        assert preview["total_validos"] == 1
        assert preview["cambios"][0]["new_score"] == 4.8
        assert preview["errores"] == []
        # Nada se guardo todavia.
        grade = grade_service.get_grade(escenario["activity_id"],
                                        escenario["student_id"])
        assert grade["score"] == 3.0


def test_importar_reporta_errores_sin_detener_el_resto(app, escenario):
    with app.app_context():
        buffer = excel_service.export_gradebook(
            _assignment(app, escenario), _period(app, escenario))
        book = load_workbook(buffer)
        sheet = book["Notas"]
        sheet.cell(row=5, column=3, value=9.9)      # nota fuera de escala
        sheet.cell(row=6, column=1, value="NOEXISTE")  # codigo inexistente
        from io import BytesIO
        modificado = BytesIO()
        book.save(modificado)
        modificado.seek(0)

        preview = excel_service.preview_import(
            modificado, escenario["assignment_id"], escenario["period_id"])
        assert len(preview["errores"]) == 2
        assert preview["total_validos"] == 0


def test_confirmar_importacion_guarda_los_cambios(app, escenario):
    with app.app_context():
        cambios = [{
            "student_id": escenario["student_id"],
            "activity_id": escenario["activity_id"],
            "new_score": 4.4,
        }]
        aplicados = excel_service.apply_import(
            cambios, escenario["assignment_id"], escenario["period_id"], USER)
        assert aplicados == 1
        grade = grade_service.get_grade(escenario["activity_id"],
                                        escenario["student_id"])
        assert grade["score"] == 4.4


def test_una_importacion_invalida_no_deja_datos_a_medias(app, escenario):
    """El segundo cambio es invalido: no debe aplicarse ninguno."""
    with app.app_context():
        cambios = [
            {"student_id": escenario["student_id"],
             "activity_id": escenario["activity_id"], "new_score": 4.9},
            {"student_id": escenario["student2_id"],
             "activity_id": escenario["activity_id"], "new_score": 99},
        ]
        with pytest.raises(grade_service.ValidationError):
            excel_service.apply_import(cambios, escenario["assignment_id"],
                                       escenario["period_id"], USER)

        grade = grade_service.get_grade(escenario["activity_id"],
                                        escenario["student_id"])
        assert grade["score"] == 3.0   # la nota original sigue intacta


def test_no_se_pueden_importar_estudiantes_de_otro_grupo(app, escenario):
    with app.app_context():
        cambios = [{"student_id": escenario["outsider_id"],
                    "activity_id": escenario["activity_id"], "new_score": 4.0}]
        with pytest.raises(excel_service.ImportError_):
            excel_service.apply_import(cambios, escenario["assignment_id"],
                                       escenario["period_id"], USER)
