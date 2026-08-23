"""Exportacion e importacion de notas en Excel.

El objetivo es que Excel deje de ser la fuente de verdad y pase a ser un
canal de entrada, salida y respaldo. Por eso la importacion nunca escribe
directo: sube, lee, valida, previsualiza, confirma y recien entonces
guarda, en una unica transaccion.
"""
import re
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.db import query_all, query_one, transaction
from services import audit_service, grade_service

HEADER_FIRST_CELL = "Codigo"
_WEIGHT_SUFFIX = re.compile(r"\s*\(\s*[\d.,]+\s*%\s*\)\s*$")


class ImportError_(Exception):
    """Archivo con formato no reconocido."""


def _clean_activity_name(value: str) -> str:
    return _WEIGHT_SUFFIX.sub("", str(value or "")).strip()


def build_filename(assignment: dict, period: dict) -> str:
    def slug(text: str) -> str:
        text = re.sub(r"[^\w\s-]", "", str(text), flags=re.UNICODE)
        return re.sub(r"[\s-]+", "_", text.strip())

    return f"{slug(assignment['subject_name'])}_{slug(assignment['group_name'])}_" \
           f"{slug(period['name'])}.xlsx"


# ------------------------------------------------------------ EXPORTACION ---
def export_gradebook(assignment: dict, period: dict) -> BytesIO:
    book = grade_service.gradebook(assignment["id"], period["id"])
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notas"

    title = (f"AcademyNotes - {assignment['subject_name']} - "
             f"{assignment['group_name']} - {period['name']}")
    sheet.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    sheet.cell(row=2, column=1,
               value=f"asignacion_id={assignment['id']} periodo_id={period['id']} "
                     "(no modificar esta fila)")
    sheet.cell(row=2, column=1).font = Font(italic=True, size=9, color="777777")

    header_row = 4
    headers = [HEADER_FIRST_CELL, "Estudiante"]
    for activity in book["activities"]:
        headers.append(f"{activity['name']} ({activity['weight']:g}%)")
    headers.append("Promedio")

    fill = PatternFill("solid", fgColor="1F4E79")
    for index, text in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=index, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for offset, student in enumerate(book["students"], start=1):
        row = header_row + offset
        sheet.cell(row=row, column=1, value=student["student_code"])
        sheet.cell(row=row, column=2, value=student["full_name"])
        for index, cell_data in enumerate(student["cells"], start=3):
            sheet.cell(row=row, column=index, value=cell_data["score"])
        sheet.cell(row=row, column=len(headers), value=student["average"])

    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 34
    for index in range(3, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(row=header_row, column=index)
                                .column_letter].width = 16
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=3)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# ------------------------------------------------------------ IMPORTACION ---
def preview_import(file_stream, assignment_id: int, period_id: int) -> dict:
    """Lee el archivo y devuelve los cambios detectados SIN guardar nada."""
    try:
        workbook = load_workbook(file_stream, data_only=True)
    except Exception as error:  # archivo corrupto o formato incorrecto
        raise ImportError_(f"No se pudo leer el archivo: {error}")

    sheet = workbook["Notas"] if "Notas" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    header_index = None
    for index, row in enumerate(rows):
        if row and str(row[0]).strip().lower() == HEADER_FIRST_CELL.lower():
            header_index = index
            break
    if header_index is None:
        raise ImportError_(
            "No se encontro la fila de encabezados. Usa el archivo exportado "
            "por AcademyNotes como plantilla."
        )

    headers = rows[header_index]
    activities = query_all(
        "SELECT id, name, weight FROM activities "
        "WHERE assignment_id = ? AND period_id = ? ORDER BY id",
        (assignment_id, period_id),
    )
    by_name = {a["name"].strip().lower(): a for a in activities}

    column_map: dict[int, dict] = {}
    errors: list[dict] = []
    for column_index, header in enumerate(headers):
        if column_index < 2 or header is None:
            continue
        name = _clean_activity_name(header)
        if not name or name.lower() == "promedio":
            continue
        activity = by_name.get(name.lower())
        if activity is None:
            errors.append({
                "fila": header_index + 1,
                "detalle": f"La columna '{name}' no corresponde a ninguna "
                           "actividad de este periodo.",
            })
            continue
        column_map[column_index] = activity

    students = query_all(
        """
        SELECT s.id, s.student_code, u.full_name
        FROM students s
        JOIN users u ON u.id = s.user_id
        JOIN teaching_assignments ta ON ta.group_id = s.group_id
        WHERE ta.id = ?
        """,
        (assignment_id,),
    )
    by_code = {str(s["student_code"]).strip().lower(): s for s in students}

    changes: list[dict] = []
    unchanged = 0
    for offset, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        code = str(row[0]).strip()
        student = by_code.get(code.lower())
        if student is None:
            errors.append({
                "fila": offset,
                "detalle": f"El codigo '{code}' no corresponde a un estudiante "
                           "de este grupo.",
            })
            continue

        for column_index, activity in column_map.items():
            raw = row[column_index] if column_index < len(row) else None
            if raw is None or str(raw).strip() == "":
                continue
            try:
                value = grade_service.validate_score(raw)
            except grade_service.ValidationError as error:
                errors.append({
                    "fila": offset,
                    "detalle": f"{student['full_name']} / {activity['name']}: {error}",
                })
                continue

            current = grade_service.get_grade(activity["id"], student["id"])
            old = current["score"] if current else None
            if old is not None and abs(float(old) - float(value)) < 1e-9:
                unchanged += 1
                continue
            changes.append({
                "student_id": student["id"],
                "student_code": student["student_code"],
                "student_name": student["full_name"],
                "activity_id": activity["id"],
                "activity_name": activity["name"],
                "old_score": old,
                "new_score": value,
                "accion": "Nueva" if old is None else "Actualiza",
            })

    return {
        "assignment_id": assignment_id,
        "period_id": period_id,
        "cambios": changes,
        "errores": errors,
        "sin_cambios": unchanged,
        "total_validos": len(changes),
    }


def apply_import(changes: list[dict], assignment_id: int, period_id: int,
                 user: dict) -> int:
    """Aplica los cambios confirmados en una sola transaccion.

    Si algo falla a mitad de camino no queda nada aplicado.
    """
    valid_activities = {
        row["id"] for row in query_all(
            "SELECT id FROM activities WHERE assignment_id = ? AND period_id = ?",
            (assignment_id, period_id),
        )
    }
    valid_students = {
        row["id"] for row in query_all(
            """
            SELECT s.id FROM students s
            JOIN teaching_assignments ta ON ta.group_id = s.group_id
            WHERE ta.id = ?
            """,
            (assignment_id,),
        )
    }

    prepared = []
    for change in changes:
        activity_id = int(change["activity_id"])
        student_id = int(change["student_id"])
        if activity_id not in valid_activities:
            raise ImportError_("La importacion incluye una actividad ajena a esta asignacion.")
        if student_id not in valid_students:
            raise ImportError_("La importacion incluye un estudiante ajeno a este grupo.")
        score = grade_service.validate_score(change["new_score"])
        current = grade_service.get_grade(activity_id, student_id)
        prepared.append((activity_id, student_id, score,
                         current["score"] if current else None))

    with transaction() as connection:
        for activity_id, student_id, score, _old in prepared:
            connection.execute(
                """
                INSERT INTO grades (activity_id, student_id, score, updated_by)
                VALUES (?,?,?,?)
                ON CONFLICT(activity_id, student_id) DO UPDATE SET
                    score = excluded.score,
                    updated_by = excluded.updated_by,
                    updated_at = datetime('now','localtime')
                """,
                (activity_id, student_id, score, user["id"]),
            )

    assignment = query_one(
        """
        SELECT s.name AS subject_name, g.name AS group_name
        FROM teaching_assignments ta
        JOIN subjects s ON s.id = ta.subject_id
        JOIN student_groups g ON g.id = ta.group_id
        WHERE ta.id = ?
        """,
        (assignment_id,),
    )
    audit_service.log(
        user, "Importo notas desde Excel", "asignacion", assignment_id,
        f"{assignment['subject_name']} - {assignment['group_name']}: "
        f"{len(prepared)} calificaciones actualizadas",
        None, len(prepared),
    )
    return len(prepared)
